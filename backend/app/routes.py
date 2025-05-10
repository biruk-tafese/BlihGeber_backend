import bcrypt

from flask import Blueprint, request, jsonify


from .userModel import User
from .db import db
import jwt
import datetime
import joblib
import pandas as pd
import os
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from flask import request
from matplotlib.backends.backend_pdf import PdfPages
from fpdf import FPDF
import pandas as pd
from io import BytesIO
from flask import send_file
from datetime import datetime
import csv
from openpyxl import Workbook
import io

# from .__init__ import create_app
# app = create_app()

current_dir = os.path.dirname(os.path.abspath(__file__))

# Load model and expected feature columns
model = joblib.load(os.path.join(current_dir, 'Random_Forest_model.pkl'))
expected_columns = joblib.load(os.path.join(current_dir, 'model_features.pkl'))


# Dynamically extract crops and areas
crop_prefix = "Item_"
area_prefix = "Area_"
crops = sorted([col[len(crop_prefix):] for col in expected_columns if col.startswith(crop_prefix)])
areas = sorted([col[len(area_prefix):] for col in expected_columns if col.startswith(area_prefix)])


routes = Blueprint('routes', __name__)
def process_prediction_data(request):
    """Common data processing for all formats"""
    data = request.json
    
    # Extract and process data (same as before)
    avg_rain = data.get('average_rain_fall_mm_per_year', 0.0)
    pesticides = data.get('pesticides_tonnes', 0.0)
    avg_temp = data.get('avg_temp', 0.0)
    selected_crop = data.get('crop', '')
    selected_area = data.get('area', '')


    # Build input dict
    input_data = {
        'average_rain_fall_mm_per_year': [avg_rain],
        'pesticides_tonnes': [pesticides],
        'avg_temp': [avg_temp],
    }

    # One-hot encode crop and area
    for crop in crops:
        input_data[f'Item_{crop}'] = [1 if crop == selected_crop else 0]
    for area in areas:
        input_data[f'Area_{area}'] = [1 if area == selected_area else 0]

    # Convert to DataFrame
    input_df = pd.DataFrame(input_data)

    # Reindex to match expected columns (fill missing with 0)
    input_df = input_df.reindex(columns=expected_columns, fill_value=0)
    
    # Return formatted data
    return {
        'Crop': selected_crop,
        'Area': selected_area,
        'Average Rainfall (mm/yr)': avg_rain,
        'Pesticides (tonnes)': pesticides,
        'Average Temperature (°C)': avg_temp,
        'Predicted Yield (tons/hectare)':input_df,
        'Generated At': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

"""
Decorator function to enforce JWT token authentication for protected routes.

This decorator checks for the presence of a valid JWT token in the `Authorization` 
header of the incoming request. If the token is missing, expired, or invalid, 
it returns an appropriate error response. If the token is valid, it retrieves 
the current user from the database and passes it as the first argument to the 
decorated function.

Args:
    f (function): The function to be decorated.

Returns:
    function: The decorated function with token authentication applied.

Raises:
    401 Unauthorized: If the token is missing, expired, or invalid.

Example:
    @token_required
    def protected_route(current_user):
        return jsonify({'message': 'This is a protected route'})
"""
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        if 'Authorization' in request.headers:
            bearer = request.headers['Authorization']
            token = bearer.split()[1] if ' ' in bearer else bearer

        if not token:
            return jsonify({'error': 'Token is missing'}), 401

        try:
            data = jwt.decode(token, 'AAiT CropYieldPrediction Project', algorithms=['HS256'])
            current_user = User.query.get(data['user_id'])
        except jwt.ExpiredSignatureError:
            return jsonify({'error': 'Token has expired'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'error': 'Invalid token'}), 401

        return f(current_user, *args, **kwargs)
    return decorated


 # Register Endpoint
@routes.route('/register', methods=['POST'])
def register_user():
    data = request.get_json()

    if not all(k in data for k in ('phone_number', 'full_name', 'password')):
        return jsonify({'error': 'Missing fields'}), 400

    if User.query.filter_by(phone_number=data['phone_number']).first():
        return jsonify({'error': 'This Account already exists'}), 409

    hashed_password = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt())

    new_user = User(
        full_name=data['full_name'],
        phone_number=data['phone_number'],
        password=hashed_password.decode('utf-8')
    )
    db.session.add(new_user)
    db.session.commit()

    return jsonify({'message': 'User registered'}), 201


# Login Endpoint with JWT
@routes.route('/login', methods=['POST'])
def login_user():
    data = request.get_json()

    if not all(k in data for k in ('phone_number', 'password')):
        return jsonify({'error': 'Missing phone number or password'}), 400

    user = User.query.filter_by(phone_number=data['phone_number']).first()
    if not user or not bcrypt.checkpw(data['password'].encode('utf-8'), user.password.encode('utf-8')):
        return jsonify({'error': 'Invalid credentials'}), 401

    # Generate JWT token
    token = jwt.encode({
        'user_id': user.id,
        # 'exp': datetime.now() + datetime.timedelta(hours=2)  # Token expires in 2 hours
    }, 'AAiT CropYieldPrediction Project', algorithm='HS256')

    return jsonify({'message': 'Login successful', 'token': token, 'user_type': user.user_type}), 200

@routes.route('/crop', methods=['GET'])
@token_required
def get_crops(current_user):
    return jsonify(crops)
@routes.route('/area', methods=['GET'])
@token_required
def get_areas(current_user):
    return jsonify(areas)
@routes.route('/')
def index():
    return jsonify("Welcome to the Crop Yield Prediction API! Use /predict to make predictions.")

@routes.route('/predict', methods=['POST'])
@token_required
def predict():
    data = request.json

    # Extract input values
    avg_rain = data.get('average_rain_fall_mm_per_year', 0.0)
    pesticides = data.get('pesticides_tonnes', 0.0)
    avg_temp = data.get('avg_temp', 0.0)
    selected_crop = data.get('crop', '')
    selected_area = data.get('area', '')

    # Build input dict
    input_data = {
        'average_rain_fall_mm_per_year': [avg_rain],
        'pesticides_tonnes': [pesticides],
        'avg_temp': [avg_temp],
    }

    # One-hot encode crop and area
    for crop in crops:
        input_data[f'Item_{crop}'] = [1 if crop == selected_crop else 0]
    for area in areas:
        input_data[f'Area_{area}'] = [1 if area == selected_area else 0]

    # Convert to DataFrame
    input_df = pd.DataFrame(input_data)

    # Reindex to match expected columns (fill missing with 0)
    input_df = input_df.reindex(columns=expected_columns, fill_value=0)

    # Make prediction
    prediction = model.predict(input_df)
    return jsonify({'predicted_yield': prediction[0]})


@routes.route('/predict/download-result-pdf/', methods=['POST'])
# @token_required
def predict_download_result():
    data = request.json

    avg_rain = data.get('average_rain_fall_mm_per_year', 0.0)
    pesticides = data.get('pesticides_tonnes', 0.0)
    avg_temp = data.get('avg_temp', 0.0)
    selected_crop = data.get('crop', '')
    selected_area = data.get('area', '')

    input_data = {
        'average_rain_fall_mm_per_year': [avg_rain],
        'pesticides_tonnes': [pesticides],
        'avg_temp': [avg_temp],
    }

    for crop in crops:
        input_data[f'Item_{crop}'] = [1 if crop == selected_crop else 0]
    for area in areas:
        input_data[f'Area_{area}'] = [1 if area == selected_area else 0]

    input_df = pd.DataFrame(input_data)
    input_df = input_df.reindex(columns=expected_columns, fill_value=0)

    prediction = model.predict(input_df)[0]

    # Create PDF with improved styling
    pdf = FPDF()
    pdf.add_page()
    
    # Add custom fonts and colors
    primary_color = (0, 102, 204)  # Blue
    accent_color = (76, 175, 80)   # Green
    text_color = (51, 51, 51)      # Dark gray
    
    # Header Section
    pdf.set_fill_color(*primary_color)
    pdf.rect(0, 0, 210, 40, style='F')
    pdf.set_font("Arial", 'B', 24)
    pdf.set_text_color(255, 255, 255)
    pdf.set_xy(10, 12)
    pdf.cell(0, 10, "Crop Yield Prediction Report", 0, 1)
    
    # Subheader
    pdf.set_font("Arial", 'I', 12)
    pdf.set_xy(10, 25)
    pdf.cell(0, 10, "Smart Agricultural Analysis Report", 0, 1)
    
    # Reset text color
    pdf.set_text_color(*text_color)
    
    # User Input Section
    pdf.set_font("Arial", 'B', 14)
    pdf.set_y(45)
    pdf.cell(0, 10, "Input Parameters", 0, 1)
    pdf.set_line_width(0.5)
    pdf.set_draw_color(*primary_color)
    pdf.line(10, 55, 200, 55)
    
    # Input Parameters Table
    pdf.set_font("Arial", size=12)
    data = [
        ("Crop Selected", selected_crop),
        ("Area", selected_area),
        ("Average Rainfall (mm/yr)", f"{avg_rain:.2f}"),
        ("Pesticides (tonnes)", f"{pesticides:.2f}"),
        ("Average Temperature (°C)", f"{avg_temp:.2f}")
    ]
    
    y_position = 60
    for label, value in data:
        pdf.set_xy(15, y_position)
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(50, 8, label)
        pdf.set_xy(65, y_position)
        pdf.set_font("Arial", '', 12)
        pdf.cell(0, 8, str(value))
        y_position += 8
    
    # Prediction Result Section
    pdf.set_font("Arial", 'B', 14)
    pdf.set_y(y_position + 10)
    pdf.cell(0, 10, "Prediction Result", 0, 1)
    pdf.line(10, y_position + 20, 200, y_position + 20)
    
    # Highlighted Prediction
    pdf.set_xy(10, y_position + 25)
    pdf.set_fill_color(*accent_color)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(0, 12, f" Predicted Yield: {prediction:.2f} tons/hectare ", 0, 1, fill=True)
    
    # Additional Information
    pdf.set_text_color(*text_color)
    pdf.set_font("Arial", 'I', 10)
    pdf.set_y(260)
    pdf.cell(0, 10, f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 0, 0, 'C')
    
    # Create PDF in-memory
    pdf_output = BytesIO()
    pdf_output.write(pdf.output(dest='S').encode('latin-1'))
    pdf_output.seek(0)

    return send_file(pdf_output, as_attachment=True, download_name="crop_prediction_report.pdf", mimetype='application/pdf')

@routes.route('/predict/download-result-csv/', methods=['POST'])
def download_csv():
    report_data = process_prediction_data(request)

    csv_output = io.BytesIO()
    text_stream = io.TextIOWrapper(csv_output, encoding='utf-8', newline='')

    csv_writer = csv.writer(text_stream)

    # Write header and data rows
    csv_writer.writerow(['Parameter', 'Value'])
    for key, value in report_data.items():
        csv_writer.writerow([key, value])

    # Flush and detach the text stream (important!)
    text_stream.flush()
    text_stream.detach()

    # Move pointer to start
    csv_output.seek(0)

    return send_file(
        csv_output,
        as_attachment=True,
        download_name=f"Crop_Prediction_{datetime.now().strftime('%Y%m%d')}.csv",
        mimetype='text/csv'
    )

@routes.route('/predict/download-result-excel/', methods=['POST'])
def download_excel():
    report_data = process_prediction_data(request)  # This must return a mix of scalars and a DataFrame

    excel_output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Crop Prediction Report"

    # Header
    ws.append(['Parameter', 'Value'])

    # Write simple key-value pairs
    for key, value in report_data.items():
        if isinstance(value, pd.DataFrame):
            # Flatten DataFrame row to individual key-value lines
            row_dict = value.iloc[0].to_dict()
            for sub_key, sub_value in row_dict.items():
                ws.append([sub_key, sub_value])
        else:
            ws.append([key, value])

    # Style columns
    for col in ['A', 'B']:
        ws.column_dimensions[col].width = 35

    wb.save(excel_output)
    excel_output.seek(0)

    return send_file(
        excel_output,
        as_attachment=True,
        download_name=f"Crop_Prediction_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
@routes.route('/profile', methods=['GET'])
@token_required
def profile(current_user):
    return jsonify({'user': current_user.to_dict()}), 200

@routes.route('/update_profile', methods=['PUT'])
@token_required
def update_profile(current_user):
    data = request.get_json()

    if not data:
        return jsonify({'error': 'No data provided'}), 400

    if 'full_name' in data:
        current_user.full_name = data['full_name']

    if 'phone_number' in data:
        current_user.phone_number = data['phone_number']

    db.session.commit()

    return jsonify({'message': 'Profile updated successfully', 'user': current_user.to_dict()}), 200




@routes.route('/update_password', methods=['PUT'])
@token_required 
def password_update(current_user):
    data = request.get_json()

    if not all(k in data for k in ('old_password', 'new_password')):
        return jsonify({'error': 'Missing fields'}), 400

    if not bcrypt.checkpw(data['old_password'].encode('utf-8'), current_user.password.encode('utf-8')):
        return jsonify({'error': 'Old password is incorrect'}), 401

    current_user.password = bcrypt.hashpw(data['new_password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    db.session.commit()

    return jsonify({'message': 'Password updated successfully'}), 200

@routes.route('/logout', methods=['POST'])
@token_required
def logout(current_user):
    # Invalidate the token (this is a simple example; in a real app, you might want to store invalidated tokens)
    return jsonify({'message': 'Logged out successfully'}), 200


@routes.route('/admin/create-user', methods=['POST'])
@token_required
def create_user(current_user):
    if current_user.user_type != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.get_json()

    if not all(k in data for k in ('phone_number', 'full_name', 'password', 'user_type')):
        return jsonify({'error': 'Missing fields'}), 400

    if User.query.filter_by(phone_number=data['phone_number']).first():
        return jsonify({'error': 'This Account already exists'}), 409

    hashed_password = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt())

    new_user = User(
        full_name=data['full_name'],
        phone_number=data['phone_number'],
        password=hashed_password.decode('utf-8'),
        user_type=data['user_type']
    )
    db.session.add(new_user)
    db.session.commit()

    return jsonify({'message': 'User created successfully'}), 201

@routes.route('/admin/update-user/<int:user_id>', methods=['PUT'])
@token_required
def update_user(current_user, user_id):
    if current_user.user_type != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.get_json()

    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'User not found'}), 404

    if 'full_name' in data:
        user.full_name = data['full_name']

    if 'phone_number' in data:
        user.phone_number = data['phone_number']


    db.session.commit()

    return jsonify({'message': 'User updated successfully', 'user': user.to_dict()}), 200   

@routes.route('/admin/delete-user/<int:user_id>', methods=['DELETE'])
@token_required
def delete_user(current_user, user_id):
    if current_user.user_type != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403

    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'User not found'}), 404

    db.session.delete(user)
    db.session.commit()

    return jsonify({'message': 'User deleted successfully'}), 200   

